# -*- coding: utf-8 -*-
"""Construit le classeur de sourcing AliExpress à partir des JSON de niche."""
import json, glob, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TMP = "/Users/Hakim/Documents/Boutiques drop/boutique-pipeline/tmp"
OUT = "/Users/Hakim/Documents/Boutiques drop/boutique-pipeline/reports/aliexpress-fournisseurs-2026-07-16.xlsx"

FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name=FONT, size=10)
LINK_FONT = Font(name=FONT, size=10, color="0563C1", underline="single")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Couleurs par décision
DEC_FILL = {
    "À APPROFONDIR":        PatternFill("solid", fgColor="C6EFCE"),
    "ALTERNATIVE":          PatternFill("solid", fgColor="DDEBF7"),
    "TROP CHER":            PatternFill("solid", fgColor="FFEB9C"),
    "RISQUE SAV":           PatternFill("solid", fgColor="FCE4D6"),
    "DONNÉES INSUFFISANTES":PatternFill("solid", fgColor="E7E6E6"),
    "À ÉCARTER":            PatternFill("solid", fgColor="FFC7CE"),
}
DEC_FONT = {
    "À APPROFONDIR":        Font(name=FONT, size=10, bold=True, color="006100"),
    "ALTERNATIVE":          Font(name=FONT, size=10, color="1F4E78"),
    "TROP CHER":            Font(name=FONT, size=10, color="9C6500"),
    "RISQUE SAV":           Font(name=FONT, size=10, color="833C0C"),
    "DONNÉES INSUFFISANTES":Font(name=FONT, size=10, color="595959"),
    "À ÉCARTER":            Font(name=FONT, size=10, color="9C0006"),
}

def norm_dec(v):
    """Canonise la décision (accents/casse variables selon les agents)."""
    s = str(v).strip().upper()
    repl = {
        "A APPROFONDIR": "À APPROFONDIR", "À APPROFONDIR": "À APPROFONDIR",
        "DONNEES INSUFFISANTES": "DONNÉES INSUFFISANTES", "DONNÉES INSUFFISANTES": "DONNÉES INSUFFISANTES",
        "A ECARTER": "À ÉCARTER", "À ÉCARTER": "À ÉCARTER", "A ÉCARTER": "À ÉCARTER",
        "TROP CHER": "TROP CHER", "ALTERNATIVE": "ALTERNATIVE", "RISQUE SAV": "RISQUE SAV",
    }
    return repl.get(s, str(v).strip())

def num(v):
    """Extrait un float d'une chaîne de prix, sinon None."""
    if v is None: return None
    s = str(v).replace(" ", "").replace("\xa0", " ").strip()
    s = s.replace("EUR", "").replace("€", "").strip()
    m = re.search(r"(\d+(?:[.,]\d+)?)", s.replace(" ", ""))
    if not m: return None
    try: return float(m.group(1).replace(",", "."))
    except ValueError: return None

# ---------- Chargement ----------
niches = []
for f in sorted(glob.glob(os.path.join(TMP, "niche-*.json"))):
    try:
        d = json.load(open(f, encoding="utf-8"))
    except Exception as e:
        print("SKIP (JSON invalide):", os.path.basename(f), e); continue
    for n in d.get("niches", []):
        n["_src"] = os.path.basename(f)
        n["_blocages"] = d.get("blocages", "")
        niches.append(n)

print(f"{len(niches)} niches chargées")

wb = Workbook()

# ================= ONGLET 1 — Fournisseurs =================
ws = wb.active
ws.title = "Fournisseurs"
COLS = [
    ("Priorité", 12), ("Niche", 26), ("Produit", 46), ("Variante analysée", 34),
    ("URL AliExpress", 30), ("Boutique", 24), ("Ancienneté boutique", 16),
    ("Note boutique", 16), ("Abonnés", 10), ("Nombre de ventes", 14),
    ("Note produit", 11), ("Nombre d'avis", 12), ("Avis avec photos", 14),
    ("Prix produit", 12), ("Livraison", 14), ("Coût total rendu", 15),
    ("TVA", 24), ("Expédition depuis", 15), ("Délai France", 22),
    ("Retour gratuit", 14), ("Badge Choice", 12), ("Garantie", 14),
    ("Prise UE", 20), ("Langue française", 16), ("CE/RoHS annoncé", 22),
    ("Caractéristiques clés", 44), ("Avis négatifs récurrents", 40),
    ("Risque SAV", 34), ("Score", 8), ("Décision", 20), ("Notes", 48),
]
for i, (h, w) in enumerate(COLS, 1):
    c = ws.cell(1, i, h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30

KEYS = ["produit","variante","url","boutique","anciennete_boutique","note_boutique","abonnes",
        "ventes","note_produit","nb_avis","avis_photos","prix","livraison","cout_rendu","tva",
        "expedie_depuis","delai_france","retour_gratuit","badge_choice","garantie","prise_ue",
        "langue_fr","ce_rohs","caracteristiques","avis_negatifs","risque_sav","score","decision","notes"]

r = 2
for n in niches:
    for f in n.get("retenus", []):
        ws.cell(r, 1, n.get("priorite", "non indiqué"))
        ws.cell(r, 2, n.get("niche", ""))
        for j, k in enumerate(KEYS, 3):
            v = f.get(k, "non indiqué")
            if k == "url":
                c = ws.cell(r, j, "Ouvrir la fiche")
                if v and str(v).startswith("http"):
                    c.hyperlink = v; c.font = LINK_FONT
                else:
                    c.value = "non indiqué"; c.font = CELL_FONT
                continue
            if k in ("prix", "cout_rendu"):
                x = num(v)
                c = ws.cell(r, j, x if x is not None else str(v))
                if x is not None: c.number_format = '#,##0.00 €'
                c.font = CELL_FONT; continue
            if k == "score":
                try: c = ws.cell(r, j, int(v))
                except (TypeError, ValueError): c = ws.cell(r, j, str(v))
                c.number_format = "0"; c.font = CELL_FONT
                c.alignment = Alignment(horizontal="center"); continue
            if k == "decision":
                dv = norm_dec(v)
                c = ws.cell(r, j, dv)
                c.fill = DEC_FILL.get(dv, PatternFill())
                c.font = DEC_FONT.get(dv, CELL_FONT)
                c.alignment = Alignment(horizontal="center", wrap_text=True); continue
            c = ws.cell(r, j, str(v)); c.font = CELL_FONT
        for j in range(1, len(COLS) + 1):
            ws.cell(r, j).border = BORDER
            al = ws.cell(r, j).alignment
            ws.cell(r, j).alignment = Alignment(vertical="top", wrap_text=True,
                                               horizontal=al.horizontal)
        r += 1
NB_FOURN = r - 2
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{max(r-1,1)}"

# ================= ONGLET 2 — Comparatif =================
ws2 = wb.create_sheet("Comparatif")
C2 = [("Niche", 30), ("Meilleur fournisseur", 30), ("Deuxième choix", 30),
      ("Meilleur coût rendu", 17), ("Prix de vente envisagé", 19),
      ("Écart brut *", 14), ("Délai", 22), ("Principal avantage", 46),
      ("Principal risque", 52), ("Décision provisoire", 20)]
for i, (h, w) in enumerate(C2, 1):
    c = ws2.cell(1, i, h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 30

r2 = 2
for n in niches:
    ret = sorted(n.get("retenus", []), key=lambda x: -(x.get("score") or 0))
    if not ret: continue
    b = ret[0]; second = ret[1]["boutique"] if len(ret) > 1 else "aucun"
    cr = num(b.get("cout_rendu"))
    pv = n.get("prix_vente_envisage", "non indiqué")
    pvmin = num(pv.split("-")[0]) if pv and pv != "non indiqué" else None
    ws2.cell(r2, 1, n.get("niche", ""))
    ws2.cell(r2, 2, f'{b.get("boutique","")} — {b.get("produit","")[:40]}')
    ws2.cell(r2, 3, second)
    c = ws2.cell(r2, 4, cr if cr is not None else "non indiqué")
    if cr is not None: c.number_format = '#,##0.00 €'
    ws2.cell(r2, 5, pv)
    if cr is not None and pvmin is not None:
        c = ws2.cell(r2, 6, f"=E{r2 and pvmin and 0 or 0}")  # placeholder remplacé ci-dessous
        ws2.cell(r2, 6).value = round(pvmin - cr, 2)
        ws2.cell(r2, 6).number_format = '#,##0.00 €'
    else:
        ws2.cell(r2, 6, "non calculable")
    ws2.cell(r2, 7, b.get("delai_france", "non indiqué"))
    ws2.cell(r2, 8, (b.get("notes") or "")[:300])
    ws2.cell(r2, 9, (b.get("risque_sav") or "")[:300])
    dv = norm_dec(b.get("decision", ""))
    c = ws2.cell(r2, 10, dv)
    c.fill = DEC_FILL.get(dv, PatternFill())
    c.font = DEC_FONT.get(dv, CELL_FONT)
    for j in range(1, len(C2) + 1):
        cell = ws2.cell(r2, j); cell.border = BORDER
        if j != 10:                      # col 10 (décision) garde sa police DEC_FONT
            cell.font = CELL_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    r2 += 1

ws2.cell(r2 + 1, 1, "* Écart brut avant TVA, paiement, livraison client, retours, SAV et publicité. "
                    "Ce n'est PAS une marge nette.").font = Font(name=FONT, size=9, italic=True, color="9C0006")
ws2.freeze_panes = "B2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(C2))}{max(r2-1,1)}"

# ================= ONGLET 3 — Rejets =================
ws3 = wb.create_sheet("Rejets")
C3 = [("Niche", 30), ("Produit", 56), ("URL", 26), ("Motif du rejet", 34),
      ("Prix réel", 14), ("Problème observé", 74)]
for i, (h, w) in enumerate(C3, 1):
    c = ws3.cell(1, i, h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[1].height = 30

r3 = 2
for n in niches:
    for x in n.get("rejets", []):
        ws3.cell(r3, 1, n.get("niche", ""))
        ws3.cell(r3, 2, str(x.get("produit", "")))
        u = x.get("url", "")
        c = ws3.cell(r3, 3, "Ouvrir la fiche")
        if u and str(u).startswith("http"): c.hyperlink = u; c.font = LINK_FONT
        else: c.value = str(u or "non indiqué"); c.font = CELL_FONT
        ws3.cell(r3, 4, str(x.get("motif", "")))
        p = num(x.get("prix_reel"))
        c = ws3.cell(r3, 5, p if p is not None else str(x.get("prix_reel", "non indiqué")))
        if p is not None: c.number_format = '#,##0.00 €'
        ws3.cell(r3, 6, str(x.get("probleme", "")))
        for j in range(1, len(C3) + 1):
            cell = ws3.cell(r3, j); cell.border = BORDER
            if j != 3: cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        r3 += 1
NB_REJ = r3 - 2
ws3.freeze_panes = "B2"
ws3.auto_filter.ref = f"A1:{get_column_letter(len(C3))}{max(r3-1,1)}"

# ================= ONGLET 4 — Contrôles manuels =================
ws4 = wb.create_sheet("Contrôles manuels")
C4 = [("Fournisseur finaliste", 34), ("Niche", 26), ("Contrôle à effectuer", 44),
      ("Fait ?", 9), ("Résultat / réponse fournisseur", 46), ("Date", 12)]
for i, (h, w) in enumerate(C4, 1):
    c = ws4.cell(1, i, h); c.fill = HDR_FILL; c.font = HDR_FONT
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.row_dimensions[1].height = 30

CHECKS = [
    "Demander la déclaration UE de conformité",
    "Vérifier CE/RoHS (documents, pas logo)",
    "Vérifier l'identité du fabricant",
    "Confirmer la garantie",
    "Confirmer l'adresse de retour (UE ?)",
    "Vérifier la prise UE",
    "Vérifier la notice française",
    "Confirmer le délai réel",
    "Demander une facture",
    "Commander un échantillon",
    "Tester le produit",
    "Vérifier les médias et droits d'utilisation",
    "Faire lever la limite de quantité par commande",
    "Revalider le prix (offres « Deal du Jour » expirées)",
]
YELLOW = PatternFill("solid", fgColor="FFFF00")
r4 = 2
for n in niches:
    ret = sorted(n.get("retenus", []), key=lambda x: -(x.get("score") or 0))
    for b in ret[:2]:  # finalistes = 2 meilleurs par niche
        start = r4
        for chk in CHECKS:
            ws4.cell(r4, 1, b.get("boutique", ""))
            ws4.cell(r4, 2, n.get("niche", ""))
            ws4.cell(r4, 3, chk)
            ws4.cell(r4, 4, "").fill = YELLOW
            ws4.cell(r4, 5, "").fill = YELLOW
            ws4.cell(r4, 6, "").fill = YELLOW
            for j in range(1, len(C4) + 1):
                cell = ws4.cell(r4, j); cell.border = BORDER; cell.font = CELL_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            r4 += 1
ws4.freeze_panes = "C2"
ws4.auto_filter.ref = f"A1:{get_column_letter(len(C4))}{max(r4-1,1)}"
ws4.cell(r4 + 1, 1, "Légende : les cellules JAUNES sont à remplir par vous (Fait ? / Résultat / Date). "
                    "Le reste est généré.").font = Font(name=FONT, size=9, italic=True)

wb.save(OUT)
print(f"OK -> {OUT}")
print(f"Fournisseurs: {NB_FOURN} | Rejets: {NB_REJ} | Contrôles: {r4-2} lignes")
